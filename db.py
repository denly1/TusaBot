import os
import asyncpg
from typing import Optional, Any, Dict
from dotenv import load_dotenv

# Загружаем переменные окружения из .env файла
load_dotenv()

# Environment variables for database connection
DB_HOST = os.getenv("DB_HOST", "127.0.0.1")
DB_PORT = int(os.getenv("DB_PORT", "5432"))
DB_NAME = os.getenv("DB_NAME", "largent")
DB_USER = os.getenv("DB_USER", "tusabot_user")  # Исправлен дефолт
DB_PASSWORD = os.getenv("DB_PASSWORD", "1")


async def create_pool() -> asyncpg.Pool:
    return await asyncpg.create_pool(
        host=DB_HOST,
        port=DB_PORT,
        database=DB_NAME,
        user=DB_USER,
        password=DB_PASSWORD,
        min_size=1,
        max_size=10,
        command_timeout=30,
    )


async def init_schema(pool: asyncpg.Pool) -> None:
    async with pool.acquire() as conn:
        await conn.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                tg_id BIGINT PRIMARY KEY,
                name TEXT,
                gender TEXT CHECK (gender IN ('male', 'female')),
                age INTEGER CHECK (age >= 14 AND age <= 100),
                vk_id TEXT,
                username TEXT,
                registered_at TIMESTAMPTZ DEFAULT now(),
                created_at TIMESTAMPTZ DEFAULT now(),
                updated_at TIMESTAMPTZ DEFAULT now()
            );
            """
        )
        
        # Создание функции для автоматического обновления updated_at
        await conn.execute(
            """
            CREATE OR REPLACE FUNCTION update_updated_at_column()
            RETURNS TRIGGER AS $$
            BEGIN
                NEW.updated_at = now();
                RETURN NEW;
            END;
            $$ language 'plpgsql';
            """
        )
        
        # Создание триггера для автоматического обновления updated_at в таблице users
        await conn.execute(
            """
            DROP TRIGGER IF EXISTS update_users_updated_at ON users;
            """
        )
        await conn.execute(
            """
            CREATE TRIGGER update_users_updated_at 
                BEFORE UPDATE ON users 
                FOR EACH ROW 
                EXECUTE FUNCTION update_updated_at_column();
            """
        )


async def upsert_user(
    pool: asyncpg.Pool,
    tg_id: int,
    name: Optional[str] = None,
    gender: Optional[str] = None,
    age: Optional[int] = None,
    vk_id: Optional[str] = None,
    username: Optional[str] = None,
) -> None:
    import logging
    logger = logging.getLogger("TusaBot")
    
    async with pool.acquire() as conn:
        try:
            logger.info("Upserting user %s: name=%s, gender=%s, age=%s, vk_id=%s, username=%s", 
                       tg_id, name, gender, age, vk_id, username)
            await conn.execute(
                """
                INSERT INTO users (tg_id, name, gender, age, vk_id, username)
                VALUES ($1, $2, $3, $4, $5, $6)
                ON CONFLICT (tg_id) DO UPDATE
                SET name = COALESCE(EXCLUDED.name, users.name),
                    gender = COALESCE(EXCLUDED.gender, users.gender),
                    age = COALESCE(EXCLUDED.age, users.age),
                    vk_id = COALESCE(EXCLUDED.vk_id, users.vk_id),
                    username = COALESCE(EXCLUDED.username, users.username);
                """,
                tg_id,
                name,
                gender,
                age,
                vk_id,
                username,
            )
            logger.info("Successfully upserted user %s", tg_id)
        except Exception as e:
            logger.error("Failed to upsert user %s: %s", tg_id, e)
            raise


async def set_vk_id(pool: asyncpg.Pool, tg_id: int, vk_id: str) -> None:
    async with pool.acquire() as conn:
        await conn.execute(
            "UPDATE users SET vk_id=$2 WHERE tg_id=$1",
            tg_id,
            vk_id,
        )


async def get_user(pool: asyncpg.Pool, tg_id: int) -> Optional[Dict[str, Any]]:
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM users WHERE tg_id=$1", tg_id)
        return dict(row) if row else None


async def get_user_by_username(pool: asyncpg.Pool, username: str) -> Optional[Dict[str, Any]]:
    """Поиск пользователя по Telegram username"""
    async with pool.acquire() as conn:
        row = await conn.fetchrow("SELECT * FROM users WHERE LOWER(username)=LOWER($1)", username)
        return dict(row) if row else None


async def get_all_user_ids(pool: asyncpg.Pool) -> list[int]:
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT tg_id FROM users")
        return [r[0] for r in rows]


async def load_user_vk_data(pool: asyncpg.Pool) -> dict[int, str]:
    """Загрузить VK ID всех пользователей для кеширования"""
    async with pool.acquire() as conn:
        rows = await conn.fetch("SELECT tg_id, vk_id FROM users WHERE vk_id IS NOT NULL")
        return {row[0]: row[1] for row in rows}


async def get_user_stats(pool: asyncpg.Pool) -> dict:
    """Получить статистику пользователей"""
    async with pool.acquire() as conn:
        stats = await conn.fetchrow("""
            SELECT 
                COUNT(*) as total_users,
                COUNT(vk_id) as users_with_vk,
                COUNT(CASE WHEN gender = 'male' THEN 1 END) as male_users,
                COUNT(CASE WHEN gender = 'female' THEN 1 END) as female_users,
                COUNT(CASE WHEN registered_at >= CURRENT_DATE THEN 1 END) as today_registrations
            FROM users
        """)
        return dict(stats) if stats else {}


async def export_users_to_excel(pool: asyncpg.Pool, filename: str = "users_export.xlsx") -> str:
    """Экспорт всех пользователей в Excel файл"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from datetime import datetime
        
        async with pool.acquire() as conn:
            users = await conn.fetch("""
                SELECT 
                    tg_id,
                    name,
                    gender,
                    age,
                    vk_id,
                    registered_at,
                    created_at
                FROM users 
                ORDER BY registered_at DESC
            """)
        
        # Создаем Excel файл
        wb = Workbook()
        ws = wb.active
        ws.title = "Пользователи TusaBot"
        
        # Заголовки
        headers = [
            "Telegram ID", "Имя", "Пол", "Возраст", 
            "VK ID", "Дата регистрации", "Дата создания"
        ]
        
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Данные пользователей
        for row_idx, user in enumerate(users, 2):
            ws.cell(row=row_idx, column=1, value=user['tg_id'])
            ws.cell(row=row_idx, column=2, value=user['name'] or "Не указано")
            
            # Пол на русском
            gender_map = {"male": "Мужской", "female": "Женский"}
            ws.cell(row=row_idx, column=3, value=gender_map.get(user['gender'], "Не указано"))
            
            # Возраст
            age = user['age']
            if age:
                ws.cell(row=row_idx, column=4, value=f"{age} лет")
            else:
                ws.cell(row=row_idx, column=4, value="Не указано")
            
            ws.cell(row=row_idx, column=5, value=user['vk_id'] or "Не привязан")
            
            # Даты
            if user['registered_at']:
                ws.cell(row=row_idx, column=6, value=user['registered_at'].strftime("%d.%m.%Y %H:%M"))
            else:
                ws.cell(row=row_idx, column=6, value="Не указано")
                
            if user['created_at']:
                ws.cell(row=row_idx, column=7, value=user['created_at'].strftime("%d.%m.%Y %H:%M"))
            else:
                ws.cell(row=row_idx, column=7, value="Не указано")
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Добавляем лист со статистикой
        stats_ws = wb.create_sheet("Статистика")
        stats = await get_user_stats(pool)
        
        stats_data = [
            ["Показатель", "Значение"],
            ["Всего пользователей", stats.get('total_users', 0)],
            ["С привязанным VK", stats.get('users_with_vk', 0)],
            ["Мужчин", stats.get('male_users', 0)],
            ["Женщин", stats.get('female_users', 0)],
            ["Зарегистрировано сегодня", stats.get('today_registrations', 0)],
            ["Дата экспорта", datetime.now().strftime("%d.%m.%Y %H:%M")]
        ]
        
        for row_idx, (label, value) in enumerate(stats_data, 1):
            stats_ws.cell(row=row_idx, column=1, value=label)
            stats_ws.cell(row=row_idx, column=2, value=value)
            if row_idx == 1:  # Заголовок
                stats_ws.cell(row=row_idx, column=1).font = header_font
                stats_ws.cell(row=row_idx, column=2).font = header_font
                stats_ws.cell(row=row_idx, column=1).fill = header_fill
                stats_ws.cell(row=row_idx, column=2).fill = header_fill
        
        # Автоподбор ширины для статистики
        for column in stats_ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = max_length + 2
            stats_ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем файл
        wb.save(filename)
        return filename
        
    except Exception as e:
        raise Exception(f"Ошибка экспорта в Excel: {e}")
